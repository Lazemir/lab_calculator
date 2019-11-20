from sympy import *
import openpyxl

wb = openpyxl.load_workbook(filename = 'vars.xlsx')
sheet = wb['Vars']

var_row_index = 2
quant_of_calc = 0

class Variable:
    def __init__(self, name: str, value: float, abs_error: float):
        self.name = symbols(name)
        self.value = S(value)
        self.abs_error = S(abs_error)
        self.rel_error = self.abs_error / self.value

calculated_vars = []

while sheet.cell(row = var_row_index, column = 1).value:
    vars = calculated_vars.copy()

    if sheet.cell(row = var_row_index, column = 1).value[:3].lower() == 'new':
        name = sheet.cell(row = var_row_index, column = 1).value[4:]
        quant_of_calc += 1
        formula_row_index = var_row_index
        var_row_index += 1
    else:
        while sheet.cell(row = var_row_index, column = 1).value and sheet.cell(row = var_row_index, column = 1).value[:3].lower() != 'new':
            var = Variable(sheet.cell(row = var_row_index, column = 1).value, sheet.cell(row = var_row_index, column = 2).value,
                           sheet.cell(row = var_row_index, column = 3).value)
            vars.append(var)
            var_row_index += 1

    sum_square_errors = 0

    if sheet.cell(row = formula_row_index, column = 4).value:
        formula = S(sheet.cell(row = formula_row_index, column = 4).value)
    else:
        formula = S(sheet.cell(row=2, column=5).value)

    for var in vars:
        der = diff(formula, var.name)
        sq_err = der**2 * var.abs_error**2
        sum_square_errors += sq_err

    for var in vars:
        formula = formula.subs(var.name, var.value)
        sum_square_errors = sum_square_errors.subs(var.name, var.value)

    try:
        res = Variable(name, float(formula), sum_square_errors ** 0.5)
    except TypeError:
        error = True
    else:
        error = False

    if not error:
        calculated_vars.append(res)

        sheet.cell(row= quant_of_calc + 1, column=7).value = str(res.name)
        sheet.cell(row= quant_of_calc + 1, column=8).value = float(res.value)
        sheet.cell(row= quant_of_calc + 1, column=9).value = float(res.abs_error)
        sheet.cell(row= quant_of_calc + 1, column=10).value = float(res.rel_error)

        print('Название:', str(res.name))
        print('Значение:', float(res.value))
        print('Абсолютная погрешность:', float(res.abs_error))
        print('Относительная погрешность: ', float(res.rel_error * 100), '%')
        print()


wb.save('vars.xlsx')


